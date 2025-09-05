#!/usr/bin/env python3
"""
Convert 2025 Excel road closure data to CSV using openpyxl (pandas-free)
"""

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl not available, trying alternative approach...")

import csv
import os
from datetime import datetime

def convert_excel_to_csv_openpyxl():
    """Convert using openpyxl library"""
    try:
        excel_file = "2025 test/2025Road Closures City of Charleston.xlsx"
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        print(f"✅ Loaded Excel file with {sheet.max_row} rows, {sheet.max_column} columns")
        
        # Get headers from first row
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else f"Col_{col}")
        
        print(f"📊 Headers: {headers}")
        
        # Find REASON and STREET columns
        reason_col = None
        street_col = None
        start_col = None
        
        for i, header in enumerate(headers):
            if 'REASON' in str(header).upper():
                reason_col = i + 1
            elif 'STREET' in str(header).upper():
                street_col = i + 1
            elif 'START' in str(header).upper():
                start_col = i + 1
        
        print(f"🔍 Found columns - REASON: {reason_col}, STREET: {street_col}, START: {start_col}")
        
        # Extract all data
        all_data = []
        flood_data = []
        
        for row_num in range(2, sheet.max_row + 1):  # Skip header
            row_data = []
            for col_num in range(1, len(headers) + 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                row_data.append(str(cell_value) if cell_value else "")
            
            all_data.append(row_data)
            
            # Check if this is a flood record
            if reason_col and len(row_data) >= reason_col:
                reason_value = str(row_data[reason_col-1]).upper()
                if 'FLOOD' in reason_value:
                    flood_data.append(row_data)
        
        print(f"📈 Total records: {len(all_data)}")
        print(f"🌊 Flood records: {len(flood_data)}")
        
        # Save all data
        with open("2025_road_closures_all.csv", 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(all_data)
        print("💾 Saved: 2025_road_closures_all.csv")
        
        # Save flood data  
        if flood_data:
            # Sort by START time if available
            if start_col and len(flood_data[0]) >= start_col:
                try:
                    flood_data.sort(key=lambda x: x[start_col-1])
                    print("✅ Sorted flood records by START time")
                except:
                    print("⚠️ Could not sort by START time")
            
            with open("2025_flood_records.csv", 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(flood_data)
            print("💾 Saved: 2025_flood_records.csv")
            
            # Show flood streets
            if street_col and len(flood_data[0]) >= street_col:
                streets = [row[street_col-1] for row in flood_data if row[street_col-1]]
                unique_streets = list(set(streets))
                print(f"🛣️ Unique flood streets: {len(unique_streets)}")
                print(f"   Streets: {unique_streets}")
            
            return "2025_flood_records.csv", len(flood_data)
        
    except ImportError:
        print("❌ openpyxl not available")
        return None, 0
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return None, 0

def manual_excel_extraction():
    """Manual approach - ask user to convert Excel to CSV"""
    print("📋 Manual Excel Conversion Required:")
    print("   1. Open '2025 test/2025Road Closures City of Charleston.xlsx'")
    print("   2. Save as CSV: '2025_road_closures_all.csv'")
    print("   3. Filter REASON='FLOOD' and save as: '2025_flood_records.csv'")
    print("   4. Sort by START time")
    print("   5. Re-run this script after conversion")
    
    # Check if CSV already exists
    if os.path.exists("2025_flood_records.csv"):
        print("\n✅ Found existing 2025_flood_records.csv")
        
        # Read and analyze the CSV
        with open("2025_flood_records.csv", 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader)
            rows = list(reader)
            
        print(f"📊 Headers: {headers}")
        print(f"🌊 Flood records: {len(rows)}")
        
        # Extract streets
        if 'STREET' in headers:
            street_idx = headers.index('STREET')
            streets = [row[street_idx] for row in rows if len(row) > street_idx and row[street_idx]]
            unique_streets = list(set(streets))
            print(f"🛣️ Unique streets: {len(unique_streets)}")
            print(f"   Streets: {unique_streets[:10]}...")  # Show first 10
        
        return "2025_flood_records.csv", len(rows)
    
    return None, 0

def main():
    """Main conversion function"""
    print("🔄 Converting 2025 Excel data...")
    
    # Try openpyxl first
    flood_file, count = convert_excel_to_csv_openpyxl()
    
    if not flood_file:
        # Fall back to manual approach
        flood_file, count = manual_excel_extraction()
    
    if flood_file and count > 0:
        print(f"\n🎉 Successfully prepared {count} flood records")
        print(f"📂 File ready: {flood_file}")
    else:
        print(f"\n💥 Could not prepare flood data")
    
    return flood_file, count

if __name__ == "__main__":
    main()